import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from hr_manager import HRManager


class HRManagerTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.manager = HRManager(Path(self.temp_dir.name))
        self.manager.add_employee(
            "E001",
            "Test Employee",
            "Engineering",
            "Developer",
            25.0,
            30,
        )

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_required_workbooks_are_created(self):
        self.assertTrue(self.manager.employees_file.exists())
        self.assertTrue(self.manager.hours_file.exists())
        self.assertTrue(self.manager.report_file.exists())

    def test_duplicate_employee_id_is_rejected(self):
        with self.assertRaises(ValueError):
            self.manager.add_employee("E001", "Other", "HR", "Manager", 30, 25)

    def test_hours_and_salary_are_calculated(self):
        self.manager.record_working_hours("2026-06-01", "E001", 8, "Development")
        self.manager.record_working_hours("2026-06-02", "E001", 7.5, "Testing")

        summary = self.manager.calculate_monthly_summary("2026-06")[0]
        self.assertEqual(summary["total_hours"], 15.5)
        self.assertEqual(summary["salary"], 387.5)

    def test_vacation_balance_is_tracked(self):
        self.manager.record_vacation("E001", 4)
        employee = self.manager.list_employees()[0]
        self.assertEqual(employee["vacation_used"], 4)

        summary = self.manager.calculate_monthly_summary("2026-06")[0]
        self.assertEqual(summary["vacation_left"], 26)

    def test_excess_vacation_is_rejected(self):
        with self.assertRaises(ValueError):
            self.manager.record_vacation("E001", 31)

    def test_monthly_report_has_expected_columns(self):
        self.manager.record_working_hours("2026-06-01", "E001", 8)
        report_path = self.manager.export_monthly_report("2026-06")

        workbook = load_workbook(report_path, data_only=True)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        self.assertEqual(headers[0], "Month")
        self.assertEqual(headers[-1], "Vacation Days Left")
        self.assertEqual(sheet.cell(2, 7).value, 200)
        workbook.close()


if __name__ == "__main__":
    unittest.main()
