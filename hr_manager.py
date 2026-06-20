"""Business logic and Excel storage for the HR automation project."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


EMPLOYEE_HEADERS = [
    "Employee ID",
    "Name",
    "Department",
    "Position",
    "Hourly Wage",
    "Vacation Days Total",
    "Vacation Days Used",
]

HOURS_HEADERS = [
    "Date",
    "Employee ID",
    "Employee Name",
    "Hours Worked",
    "Notes",
]

REPORT_HEADERS = [
    "Month",
    "Employee ID",
    "Employee Name",
    "Department",
    "Total Hours",
    "Hourly Wage",
    "Salary",
    "Vacation Days Used",
    "Vacation Days Left",
]


class HRManager:
    """Manage employees, working hours, vacation, and Excel reports."""

    def __init__(self, data_dir: str | Path = "data") -> None:
        self.data_dir = Path(data_dir)
        self.data_dir.mkdir(parents=True, exist_ok=True)

        self.employees_file = self.data_dir / "employees.xlsx"
        self.hours_file = self.data_dir / "working_hours.xlsx"
        self.report_file = self.data_dir / "monthly_report.xlsx"
        self._create_files_if_missing()

    def _create_files_if_missing(self) -> None:
        """Create the required Excel files with headers."""
        self._create_workbook(self.employees_file, "Employees", EMPLOYEE_HEADERS)
        self._create_workbook(self.hours_file, "Working Hours", HOURS_HEADERS)
        self._create_workbook(self.report_file, "Monthly Report", REPORT_HEADERS)

    @staticmethod
    def _create_workbook(path: Path, sheet_name: str, headers: list[str]) -> None:
        if path.exists():
            return

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        sheet.append(headers)
        HRManager._style_sheet(sheet)
        workbook.save(path)

    @staticmethod
    def _style_sheet(sheet) -> None:
        """Apply simple, professional formatting to a worksheet."""
        dark_blue = "1F4E78"
        light_blue = "D9EAF7"
        border_color = "CBD5E1"
        thin = Side(style="thin", color=border_color)

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False

        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor=dark_blue)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=Side(style="medium", color=dark_blue))

        sheet.row_dimensions[1].height = 24

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(vertical="center")
                if cell.row % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=light_blue)

        # Size each populated column from its content, with sensible limits.
        for column_number in range(1, sheet.max_column + 1):
            longest_value = max(
                len(str(sheet.cell(row_number, column_number).value or ""))
                for row_number in range(1, sheet.max_row + 1)
            )
            width = min(max(longest_value + 3, 14), 34)
            sheet.column_dimensions[get_column_letter(column_number)].width = width

    @staticmethod
    def _refresh_formatting(sheet) -> None:
        """Reapply formatting and filters after data changes."""
        HRManager._style_sheet(sheet)
        sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"

    @staticmethod
    def _clean_required(value: str, field_name: str) -> str:
        value = value.strip()
        if not value:
            raise ValueError(f"{field_name} is required.")
        return value

    @staticmethod
    def _validate_month(month: str) -> str:
        try:
            datetime.strptime(month, "%Y-%m")
        except ValueError as error:
            raise ValueError("Month must use YYYY-MM format.") from error
        return month

    def _find_employee_row(self, employee_id: str) -> int | None:
        workbook = load_workbook(self.employees_file, read_only=True)
        sheet = workbook.active
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if str(row[0]) == employee_id:
                workbook.close()
                return row_number
        workbook.close()
        return None

    def _employee_by_id(self, employee_id: str) -> dict:
        workbook = load_workbook(self.employees_file, data_only=True, read_only=True)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == employee_id:
                workbook.close()
                return {
                    "employee_id": str(row[0]),
                    "name": row[1],
                    "department": row[2],
                    "position": row[3],
                    "hourly_wage": float(row[4]),
                    "vacation_total": int(row[5]),
                    "vacation_used": int(row[6]),
                }
        workbook.close()
        raise ValueError(f"Employee ID '{employee_id}' was not found.")

    def add_employee(
        self,
        employee_id: str,
        name: str,
        department: str,
        position: str,
        hourly_wage: float,
        vacation_days_total: int,
    ) -> None:
        employee_id = self._clean_required(employee_id, "Employee ID")
        name = self._clean_required(name, "Name")
        department = self._clean_required(department, "Department")
        position = self._clean_required(position, "Position")

        if hourly_wage < 0:
            raise ValueError("Hourly wage cannot be negative.")
        if vacation_days_total < 0:
            raise ValueError("Vacation days cannot be negative.")
        if self._find_employee_row(employee_id):
            raise ValueError(f"Employee ID '{employee_id}' already exists.")

        workbook = load_workbook(self.employees_file)
        sheet = workbook.active
        sheet.append(
            [
                employee_id,
                name,
                department,
                position,
                float(hourly_wage),
                int(vacation_days_total),
                0,
            ]
        )
        sheet.cell(sheet.max_row, 5).number_format = '#,##0.00 "EUR"'
        self._refresh_formatting(sheet)
        workbook.save(self.employees_file)

    def list_employees(self) -> list[dict]:
        workbook = load_workbook(self.employees_file, data_only=True, read_only=True)
        sheet = workbook.active
        employees = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            employees.append(
                {
                    "employee_id": str(row[0]),
                    "name": row[1],
                    "department": row[2],
                    "position": row[3],
                    "hourly_wage": float(row[4]),
                    "vacation_total": int(row[5]),
                    "vacation_used": int(row[6]),
                }
            )
        workbook.close()
        return employees

    def record_working_hours(
        self,
        work_date: str,
        employee_id: str,
        hours_worked: float,
        notes: str = "",
    ) -> None:
        try:
            parsed_date = datetime.strptime(work_date, "%Y-%m-%d").date()
        except ValueError as error:
            raise ValueError("Date must use YYYY-MM-DD format.") from error

        if not 0 <= hours_worked <= 24:
            raise ValueError("Hours worked must be between 0 and 24.")

        employee = self._employee_by_id(employee_id.strip())
        workbook = load_workbook(self.hours_file)
        sheet = workbook.active
        sheet.append(
            [
                parsed_date,
                employee["employee_id"],
                employee["name"],
                float(hours_worked),
                notes.strip(),
            ]
        )
        sheet.cell(sheet.max_row, 1).number_format = "yyyy-mm-dd"
        sheet.cell(sheet.max_row, 4).number_format = "0.00"
        self._refresh_formatting(sheet)
        workbook.save(self.hours_file)

    def record_vacation(self, employee_id: str, days_used: int) -> None:
        if days_used <= 0:
            raise ValueError("Vacation days used must be greater than zero.")

        row_number = self._find_employee_row(employee_id.strip())
        if row_number is None:
            raise ValueError(f"Employee ID '{employee_id}' was not found.")

        workbook = load_workbook(self.employees_file)
        sheet = workbook.active
        total = int(sheet.cell(row_number, 6).value)
        already_used = int(sheet.cell(row_number, 7).value)

        if already_used + days_used > total:
            workbook.close()
            raise ValueError("Vacation request exceeds the employee's remaining balance.")

        sheet.cell(row_number, 7).value = already_used + days_used
        self._refresh_formatting(sheet)
        workbook.save(self.employees_file)

    def calculate_monthly_summary(self, month: str) -> list[dict]:
        month = self._validate_month(month)
        totals: dict[str, float] = {}

        hours_workbook = load_workbook(self.hours_file, data_only=True, read_only=True)
        hours_sheet = hours_workbook.active
        for row in hours_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            work_date = row[0]
            if isinstance(work_date, datetime):
                work_date = work_date.date()
            if work_date.strftime("%Y-%m") == month:
                employee_id = str(row[1])
                totals[employee_id] = totals.get(employee_id, 0) + float(row[3])
        hours_workbook.close()

        summary = []
        for employee in self.list_employees():
            total_hours = round(totals.get(employee["employee_id"], 0), 2)
            vacation_left = employee["vacation_total"] - employee["vacation_used"]
            summary.append(
                {
                    "month": month,
                    "employee_id": employee["employee_id"],
                    "employee_name": employee["name"],
                    "department": employee["department"],
                    "total_hours": total_hours,
                    "hourly_wage": employee["hourly_wage"],
                    "salary": round(total_hours * employee["hourly_wage"], 2),
                    "vacation_used": employee["vacation_used"],
                    "vacation_left": vacation_left,
                }
            )
        return summary

    def export_monthly_report(self, month: str) -> Path:
        summary = self.calculate_monthly_summary(month)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Monthly Report"
        sheet.append(REPORT_HEADERS)

        for row in summary:
            sheet.append(
                [
                    row["month"],
                    row["employee_id"],
                    row["employee_name"],
                    row["department"],
                    row["total_hours"],
                    row["hourly_wage"],
                    row["salary"],
                    row["vacation_used"],
                    row["vacation_left"],
                ]
            )

        for row_number in range(2, sheet.max_row + 1):
            sheet.cell(row_number, 5).number_format = "0.00"
            sheet.cell(row_number, 6).number_format = '#,##0.00 "EUR"'
            sheet.cell(row_number, 7).number_format = '#,##0.00 "EUR"'

        self._refresh_formatting(sheet)
        workbook.save(self.report_file)
        return self.report_file.resolve()
